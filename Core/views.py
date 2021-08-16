from django.shortcuts import render
from django.views import View
from django.http import HttpResponse, Http404,JsonResponse
from openpyxl import load_workbook
from django.conf import settings
import xlrd
import json 

class HomePage(View):
    def costomerDetails(self,book):
        # array for specifying which sheets to iterate through 
        wsNames = ['Customer 1','Customer 2','Customer 3','Customer 4', ]
        theHolyGrail =  {}

        for customer in wsNames:
            ws = book[customer]
            #personal information
            personalDescription = []
            for cell in ws.iter_rows(min_row=1, max_row=3, min_col=2, max_col=2, values_only=True):
                # x = type(cell)
                # ('Martyn' ,'3, Covent Garden, London', 'XXXXXXXXX' )
                for x in cell: 
                    personalDescription.append(x)
        
            # consuumption details 
            consumptionDetails = []

            for cell in ws.iter_rows(min_row=6, max_row=8, min_col=2, max_col=3, values_only=True):
                for x in cell: 
                    consumptionDetails.append(x)
      
            theHolyGrail[customer] = {'personal' :personalDescription, 'consumption':  consumptionDetails}
        return theHolyGrail
    

    def get(self, request):
        # iterating over rows
        #create a workbook with the provided file
        book = load_workbook(settings.EX_FILE)
        #create a worksheet, its auotimcally start at 0
 
        context = {'data': self.costomerDetails(book)}

        # print(context)
        # return HttpResponse(context)
        return render(request, "Core/index.html", context)

    def post(self):
        raise Http404



class MaxValue(View):
    # max value function 
    #get max value of category 
    def maxValue (self, book):
        wsNames = ['Customer 1','Customer 2','Customer 3','Customer 4' ]
        sheet2 = book['Rate Price']
        holyG_2 = {}

      
        for customer in wsNames:
            sheet = book[customer]
            # read row but for col(2,3 ) subtract and have a list of 2 value
           
           #calculating consumption rate, consumption energy and total cost
            if sheet.cell(8,1).value:
                thirdRow = str(sheet.cell(8,1).value)
                weekendRat1 = float(sheet.cell(8,2).value)
                weekendRat2 = float(sheet.cell(8,3).value)
            else:
                thirdRow = 0
                weekendRat2 = 0 
                weekendRat1 = 0
             
            dayRate1 =  float(sheet.cell(6,2).value)
            dayRate2 = float(sheet.cell(6,3).value)
            nightRate1 = float(sheet.cell(7,2).value)
            nightRate2 = float(sheet.cell(7,3).value)
            dayConsumption =  dayRate2 - dayRate1
            nightConsumption =  nightRate2 - nightRate1 


            if sheet.cell(7,1).value == 'Weekend Rate': 
                nightConsumptionRate = nightConsumption * float(sheet2.cell(4,2).value)
            elif sheet.cell(7,1).value == 'Night Rate': 
                nightConsumptionRate = nightConsumption * float(sheet2.cell(3,2).value)
            else:
                nightConsumptionRate = 0

            weekendRate = weekendRat2 - weekendRat1 

            #calculating rate based off rate name 
            if thirdRow == 'Weekend Rate':
                weekendConsumptionRate =  weekendRate * float(sheet2.cell(4,2).value)
            elif thirdRow == 'Weekend Day Rate':
                weekendConsumptionRate =  weekendRate * float(sheet2.cell(5,2).value)
            else:
                weekendConsumptionRate =  weekendRate * float(sheet2.cell(6,2).value)

            dayConsumptionRate =  dayConsumption * float(sheet2.cell(2,2).value)
            totalCost = dayConsumptionRate + nightConsumptionRate + weekendConsumptionRate


            personalDescription = []
            for cell in sheet.iter_rows(min_row=1, max_row=3, min_col=2, max_col=2, values_only=True):
                # x = type(cell)
                # ('Martyn' ,'3, Covent Garden, London', 'XXXXXXXXX' )
                for x in cell: 
                    personalDescription.append(x)

            holyG_2[customer] = [{'personal': personalDescription}, dayConsumption, nightConsumption, totalCost] 

        return holyG_2


    def get(self, request):
        #return a json response 
        book = load_workbook(settings.EX_FILE)
        #create a worksheet, its auotimcally start at 0
        context = {'data': self.maxValue(book)}
        # print (context )
  
        return JsonResponse(context)

    def post(self):
        raise Http404
   

    
        
    


